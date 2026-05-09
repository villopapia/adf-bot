"""
================================================================================
 EXCEL TRACKER  -  Simple Excel-Based Paper Trading Journal
================================================================================
 Zero-setup paper trading via an Excel file (paper_trades.xlsx).
 Sheets: Pairs, Momentum, Bear, Earnings, Shock.

 How it works:
   1. When diamond signals fire, entries are auto-logged with entry price/date.
   2. Each daily run updates "Current Price" and "Unrealized P/L" for open rows.
   3. To close a position: change Status from "OPEN" to "CLOSE" in Excel.
      Next run calculates final P/L with the closing price that day.

 Called from run_system.py alongside the existing CSV trackers.
================================================================================
"""

import os, csv, datetime
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH  = os.path.join(_SCRIPT_DIR, "paper_trades.xlsx")

# ── Sheet definitions ────────────────────────────────────────────────────────

PAIRS_HEADERS = [
    "Trade ID", "Date Opened", "Stock A", "Stock B", "Direction",
    "Entry Price A", "Entry Price B", "Shares A", "Shares B",
    "Capital Deployed", "Status",
    "Current Price A", "Current Price B", "Unrealized P/L",
    "Date Closed", "Exit Price A", "Exit Price B", "Realized P/L",
    "Hold Days", "Notes",
]

MOM_HEADERS = [
    "Trade ID", "Date Opened", "Ticker", "Direction",
    "Entry Price", "Shares", "Capital Deployed", "Status",
    "Current Price", "Unrealized P/L",
    "Date Closed", "Exit Price", "Realized P/L",
    "Hold Days", "Notes",
]

BEAR_HEADERS = [
    "Trade ID", "Date Opened", "Ticker", "Module",
    "Entry Price", "Shares", "Capital Deployed", "Status",
    "Current Price", "Unrealized P/L",
    "Date Closed", "Exit Price", "Realized P/L",
    "Hold Days", "Notes",
]

SHOCK_HEADERS = [
    "Trade ID", "Date Opened", "Ticker", "Direction",
    "Entry Price", "Shares", "Capital Deployed",
    "Shock Drop%", "VIX Spike",
    "Status",
    "Current Price", "Unrealized P/L",
    "Date Closed", "Exit Price", "Realized P/L",
    "Hold Days", "Notes",
]

EARN_HEADERS = [
    "Trade ID", "Date Opened", "Ticker", "Direction",
    "Entry Price", "Shares", "Capital Deployed",
    "Earnings Date", "Beat Rate", "EPS Trend",
    "Status",
    "Current Price", "Unrealized P/L",
    "Date Closed", "Exit Price", "Realized P/L",
    "Hold Days", "Notes",
]

# ── Styles ───────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496",
                           fill_type="solid")
_OPEN_FILL   = PatternFill(start_color="E2EFDA", end_color="E2EFDA",
                           fill_type="solid")
_CLOSED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2",
                           fill_type="solid")
_GREEN_FONT  = Font(color="006100", bold=True)
_RED_FONT    = Font(color="9C0006", bold=True)
_MONEY_FMT   = '#,##0.00'
_PCT_FMT     = '0.00%'


# ── Workbook init ────────────────────────────────────────────────────────────

def _get_workbook():
    """Load existing workbook or create a fresh one with all three sheets."""
    if os.path.exists(EXCEL_PATH):
        return load_workbook(EXCEL_PATH)

    wb = Workbook()
    # Create sheets
    sheets = {
        "Pairs":    PAIRS_HEADERS,
        "Momentum": MOM_HEADERS,
        "Bear":     BEAR_HEADERS,
        "Earnings": EARN_HEADERS,
        "Shock":    SHOCK_HEADERS,
    }
    first = True
    for name, headers in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        _write_headers(ws, headers)

    wb.save(EXCEL_PATH)
    return wb


def _write_headers(ws, headers):
    """Write styled header row."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    # Auto-width estimate
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 4, 14)
    # Freeze header row
    ws.freeze_panes = "A2"


def _ensure_sheet(wb, name, headers):
    """Make sure a sheet exists (handles upgrading old files)."""
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        _write_headers(ws, headers)
    return wb[name]


def _save(wb):
    wb.save(EXCEL_PATH)


# ── Price fetching ───────────────────────────────────────────────────────────

def _fetch_price(ticker: str) -> float | None:
    """Get latest closing price via yfinance."""
    try:
        data = yf.download(ticker, period="5d", progress=False, auto_adjust=True)
        if data is not None and len(data) > 0:
            col = data["Close"]
            val = col.iloc[-1]
            return float(val.iloc[0]) if hasattr(val, "iloc") else float(val)
    except Exception:
        pass
    return None


# ── Pairs ────────────────────────────────────────────────────────────────────

def log_pair_entry(diamond: dict) -> str:
    """Log a pairs diamond signal to the Pairs sheet. Returns trade_id."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Pairs", PAIRS_HEADERS)

    today = datetime.date.today().strftime("%Y-%m-%d")
    a     = diamond["a"]
    b     = diamond["b"]
    live  = diamond["live"]
    trade_id = f"{today.replace('-', '')}_{a}_{b}"

    price_a = round(live["price_a"], 4)
    price_b = round(live["price_b"], 4)
    half    = 500.0  # $1000 / 2 legs
    shares_a = round(half / price_a)
    shares_b = round(half / price_b)
    capital  = round(shares_a * price_a + shares_b * price_b, 2)

    row = [
        trade_id, today, a, b, live["direction"],
        price_a, price_b, shares_a, shares_b,
        capital, "OPEN",
        price_a, price_b, 0.0,   # current = entry on day 1
        "", "", "", "",          # close fields
        0, "",                   # hold days, notes
    ]
    ws.append(row)
    _style_data_row(ws, ws.max_row, "OPEN")
    _save(wb)
    return trade_id


def update_pairs_prices():
    """Update current prices and unrealized P/L for all OPEN pairs."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Pairs", PAIRS_HEADERS)
    today = datetime.date.today().strftime("%Y-%m-%d")
    closed_ids = []

    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row=row_idx, column=11).value or "").strip().upper()
        if status == "OPEN":
            a = ws.cell(row=row_idx, column=3).value
            b = ws.cell(row=row_idx, column=4).value
            direction = str(ws.cell(row=row_idx, column=5).value or "")
            entry_a = float(ws.cell(row=row_idx, column=6).value or 0)
            entry_b = float(ws.cell(row=row_idx, column=7).value or 0)
            shares_a = float(ws.cell(row=row_idx, column=8).value or 0)
            shares_b = float(ws.cell(row=row_idx, column=9).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            cur_a = _fetch_price(a)
            cur_b = _fetch_price(b)
            if cur_a is not None:
                ws.cell(row=row_idx, column=12, value=round(cur_a, 4))
            if cur_b is not None:
                ws.cell(row=row_idx, column=13, value=round(cur_b, 4))

            # Unrealized P/L: LONG spread = long A / short B
            ca = cur_a if cur_a else entry_a
            cb = cur_b if cur_b else entry_b
            if "LONG" in direction.upper():
                pnl = shares_a * (ca - entry_a) - shares_b * (cb - entry_b)
            else:
                pnl = -shares_a * (ca - entry_a) + shares_b * (cb - entry_b)
            ws.cell(row=row_idx, column=14, value=round(pnl, 2))
            _color_pnl(ws, row_idx, 14, pnl)

            # Hold days
            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                hold = (datetime.date.today() - d_open).days
                ws.cell(row=row_idx, column=19, value=hold)
            except ValueError:
                pass

        elif status == "CLOSE":
            # User marked for closing -- finalize
            a = ws.cell(row=row_idx, column=3).value
            b = ws.cell(row=row_idx, column=4).value
            direction = str(ws.cell(row=row_idx, column=5).value or "")
            entry_a = float(ws.cell(row=row_idx, column=6).value or 0)
            entry_b = float(ws.cell(row=row_idx, column=7).value or 0)
            shares_a = float(ws.cell(row=row_idx, column=8).value or 0)
            shares_b = float(ws.cell(row=row_idx, column=9).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            exit_a = _fetch_price(a) or float(ws.cell(row=row_idx, column=12).value or entry_a)
            exit_b = _fetch_price(b) or float(ws.cell(row=row_idx, column=13).value or entry_b)

            if "LONG" in direction.upper():
                pnl = shares_a * (exit_a - entry_a) - shares_b * (exit_b - entry_b)
            else:
                pnl = -shares_a * (exit_a - entry_a) + shares_b * (exit_b - entry_b)

            ws.cell(row=row_idx, column=11, value="CLOSED")
            ws.cell(row=row_idx, column=12, value=round(exit_a, 4))
            ws.cell(row=row_idx, column=13, value=round(exit_b, 4))
            ws.cell(row=row_idx, column=14, value="")  # clear unrealized
            ws.cell(row=row_idx, column=15, value=today)
            ws.cell(row=row_idx, column=16, value=round(exit_a, 4))
            ws.cell(row=row_idx, column=17, value=round(exit_b, 4))
            ws.cell(row=row_idx, column=18, value=round(pnl, 2))
            _color_pnl(ws, row_idx, 18, pnl)
            _style_data_row(ws, row_idx, "CLOSED")

            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                hold = (datetime.date.today() - d_open).days
                ws.cell(row=row_idx, column=19, value=hold)
            except ValueError:
                pass

            tid = ws.cell(row=row_idx, column=1).value
            closed_ids.append(tid)

    _save(wb)
    return closed_ids


# ── Momentum ─────────────────────────────────────────────────────────────────

def log_momentum_entry(signal: dict, vix_scale: float = 1.0) -> str:
    """Log a momentum diamond to the Momentum sheet. Returns trade_id."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Momentum", MOM_HEADERS)

    today  = datetime.date.today().strftime("%Y-%m-%d")
    ticker = signal["ticker"]
    live   = signal["live"]
    price  = float(live["price"])
    capital = 1000.0 * vix_scale
    shares  = round(capital / price, 4)
    trade_id = f"{today.replace('-', '')}_{ticker}"

    row = [
        trade_id, today, ticker, "LONG",
        round(price, 4), shares, round(capital, 2), "OPEN",
        round(price, 4), 0.0,   # current = entry on day 1
        "", "", "",              # close fields
        0, "",                   # hold days, notes
    ]
    ws.append(row)
    _style_data_row(ws, ws.max_row, "OPEN")
    _save(wb)
    return trade_id


def update_momentum_prices():
    """Update current prices and unrealized P/L for all OPEN momentum trades."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Momentum", MOM_HEADERS)
    today = datetime.date.today().strftime("%Y-%m-%d")
    closed_ids = []

    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row=row_idx, column=8).value or "").strip().upper()
        if status == "OPEN":
            ticker = ws.cell(row=row_idx, column=3).value
            entry  = float(ws.cell(row=row_idx, column=5).value or 0)
            shares = float(ws.cell(row=row_idx, column=6).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            cur = _fetch_price(ticker)
            if cur is not None:
                ws.cell(row=row_idx, column=9, value=round(cur, 4))
            c = cur if cur else entry
            pnl = round(shares * (c - entry), 2)
            ws.cell(row=row_idx, column=10, value=pnl)
            _color_pnl(ws, row_idx, 10, pnl)

            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                ws.cell(row=row_idx, column=14, value=(datetime.date.today() - d_open).days)
            except ValueError:
                pass

        elif status == "CLOSE":
            ticker = ws.cell(row=row_idx, column=3).value
            entry  = float(ws.cell(row=row_idx, column=5).value or 0)
            shares = float(ws.cell(row=row_idx, column=6).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            exit_price = _fetch_price(ticker) or float(ws.cell(row=row_idx, column=9).value or entry)
            pnl = round(shares * (exit_price - entry), 2)

            ws.cell(row=row_idx, column=8, value="CLOSED")
            ws.cell(row=row_idx, column=9, value=round(exit_price, 4))
            ws.cell(row=row_idx, column=10, value="")
            ws.cell(row=row_idx, column=11, value=today)
            ws.cell(row=row_idx, column=12, value=round(exit_price, 4))
            ws.cell(row=row_idx, column=13, value=round(pnl, 2))
            _color_pnl(ws, row_idx, 13, pnl)
            _style_data_row(ws, row_idx, "CLOSED")

            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                ws.cell(row=row_idx, column=14, value=(datetime.date.today() - d_open).days)
            except ValueError:
                pass

            closed_ids.append(ws.cell(row=row_idx, column=1).value)

    _save(wb)
    return closed_ids


# ── Bear ─────────────────────────────────────────────────────────────────────

def log_bear_entry(signal: dict, module: str, vix_scale: float = 1.0) -> str:
    """Log a bear diamond to the Bear sheet. Returns trade_id."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Bear", BEAR_HEADERS)

    today  = datetime.date.today().strftime("%Y-%m-%d")
    ticker = signal["ticker"]
    live   = signal["live"]
    price  = float(live["price"])
    capital = 1000.0 * vix_scale
    shares  = round(capital / price, 4)
    trade_id = f"{today.replace('-', '')}_{module}_{ticker}"

    row = [
        trade_id, today, ticker, module,
        round(price, 4), shares, round(capital, 2), "OPEN",
        round(price, 4), 0.0,
        "", "", "",
        0, "",
    ]
    ws.append(row)
    _style_data_row(ws, ws.max_row, "OPEN")
    _save(wb)
    return trade_id


def update_bear_prices():
    """Update current prices and unrealized P/L for all OPEN bear trades."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Bear", BEAR_HEADERS)
    today = datetime.date.today().strftime("%Y-%m-%d")
    closed_ids = []

    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row=row_idx, column=8).value or "").strip().upper()
        if status == "OPEN":
            ticker = ws.cell(row=row_idx, column=3).value
            entry  = float(ws.cell(row=row_idx, column=5).value or 0)
            shares = float(ws.cell(row=row_idx, column=6).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            cur = _fetch_price(ticker)
            if cur is not None:
                ws.cell(row=row_idx, column=9, value=round(cur, 4))
            c = cur if cur else entry
            pnl = round(shares * (c - entry), 2)
            ws.cell(row=row_idx, column=10, value=pnl)
            _color_pnl(ws, row_idx, 10, pnl)

            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                ws.cell(row=row_idx, column=14, value=(datetime.date.today() - d_open).days)
            except ValueError:
                pass

        elif status == "CLOSE":
            ticker = ws.cell(row=row_idx, column=3).value
            entry  = float(ws.cell(row=row_idx, column=5).value or 0)
            shares = float(ws.cell(row=row_idx, column=6).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            exit_price = _fetch_price(ticker) or float(ws.cell(row=row_idx, column=9).value or entry)
            pnl = round(shares * (exit_price - entry), 2)

            ws.cell(row=row_idx, column=8, value="CLOSED")
            ws.cell(row=row_idx, column=9, value=round(exit_price, 4))
            ws.cell(row=row_idx, column=10, value="")
            ws.cell(row=row_idx, column=11, value=today)
            ws.cell(row=row_idx, column=12, value=round(exit_price, 4))
            ws.cell(row=row_idx, column=13, value=round(pnl, 2))
            _color_pnl(ws, row_idx, 13, pnl)
            _style_data_row(ws, row_idx, "CLOSED")

            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                ws.cell(row=row_idx, column=14, value=(datetime.date.today() - d_open).days)
            except ValueError:
                pass

            closed_ids.append(ws.cell(row=row_idx, column=1).value)

    _save(wb)
    return closed_ids


# ── Earnings ─────────────────────────────────────────────────────────────────

def log_earn_entry(signal: dict) -> str:
    """Log an earnings diamond to the Earnings sheet. Returns trade_id."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Earnings", EARN_HEADERS)

    today    = datetime.date.today().strftime("%Y-%m-%d")
    ticker   = signal["ticker"]
    live     = signal["live"]
    price    = float(live["price"])
    capital  = 1000.0
    shares   = round(capital / price, 4)
    trade_id = f"{today.replace('-', '')}_earn_{ticker}"

    beat_rate  = round(float(live.get("beat_rate", 0)) * 100, 1)
    eps_trend  = str(live.get("eps_trend", ""))
    earn_date  = str(live.get("earnings_date", ""))

    row = [
        trade_id, today, ticker, "LONG",
        round(price, 4), shares, round(capital, 2),
        earn_date, beat_rate, eps_trend,
        "OPEN",
        round(price, 4), 0.0,   # current = entry on day 1
        "", "", "",              # close fields
        0, "",                   # hold days, notes
    ]
    ws.append(row)
    _style_data_row(ws, ws.max_row, "OPEN")
    _save(wb)
    return trade_id


def update_earn_prices():
    """Update current prices and unrealized P/L for all OPEN earnings trades."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Earnings", EARN_HEADERS)
    today = datetime.date.today().strftime("%Y-%m-%d")
    closed_ids = []

    # Column layout (1-indexed):
    #  1=Trade ID, 2=Date Opened, 3=Ticker, 4=Direction
    #  5=Entry Price, 6=Shares, 7=Capital Deployed
    #  8=Earnings Date, 9=Beat Rate, 10=EPS Trend
    #  11=Status, 12=Current Price, 13=Unrealized P/L
    #  14=Date Closed, 15=Exit Price, 16=Realized P/L
    #  17=Hold Days, 18=Notes

    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row=row_idx, column=11).value or "").strip().upper()
        if status == "OPEN":
            ticker    = ws.cell(row=row_idx, column=3).value
            entry     = float(ws.cell(row=row_idx, column=5).value or 0)
            shares    = float(ws.cell(row=row_idx, column=6).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            cur = _fetch_price(ticker)
            if cur is not None:
                ws.cell(row=row_idx, column=12, value=round(cur, 4))
            c   = cur if cur else entry
            pnl = round(shares * (c - entry), 2)
            ws.cell(row=row_idx, column=13, value=pnl)
            _color_pnl(ws, row_idx, 13, pnl)

            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                ws.cell(row=row_idx, column=17,
                        value=(datetime.date.today() - d_open).days)
            except ValueError:
                pass

        elif status == "CLOSE":
            ticker    = ws.cell(row=row_idx, column=3).value
            entry     = float(ws.cell(row=row_idx, column=5).value or 0)
            shares    = float(ws.cell(row=row_idx, column=6).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            exit_price = (_fetch_price(ticker)
                          or float(ws.cell(row=row_idx, column=12).value or entry))
            pnl = round(shares * (exit_price - entry), 2)

            ws.cell(row=row_idx, column=11, value="CLOSED")
            ws.cell(row=row_idx, column=12, value=round(exit_price, 4))
            ws.cell(row=row_idx, column=13, value="")  # clear unrealized
            ws.cell(row=row_idx, column=14, value=today)
            ws.cell(row=row_idx, column=15, value=round(exit_price, 4))
            ws.cell(row=row_idx, column=16, value=round(pnl, 2))
            _color_pnl(ws, row_idx, 16, pnl)
            _style_data_row(ws, row_idx, "CLOSED")

            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                ws.cell(row=row_idx, column=17,
                        value=(datetime.date.today() - d_open).days)
            except ValueError:
                pass

            closed_ids.append(ws.cell(row=row_idx, column=1).value)

    _save(wb)
    return closed_ids


# ── Shock ────────────────────────────────────────────────────────────────────

def log_shock_entry(signal: dict, vix_scale: float = 1.0) -> str:
    """Log a shock bounce diamond to the Shock sheet. Returns trade_id."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Shock", SHOCK_HEADERS)

    today    = datetime.date.today().strftime("%Y-%m-%d")
    ticker   = signal["ticker"]
    live     = signal["live"]
    price    = float(live["price"])
    capital  = 1000.0 * vix_scale
    shares   = round(capital / price, 4)
    trade_id = f"{today.replace('-', '')}_shock_{ticker}"

    shock_drop = round(float(live.get("shock_return", 0)) * 100, 1)
    vix_spike  = round(float(live.get("vix_spike", 0)), 1)

    # Columns: Trade ID, Date Opened, Ticker, Direction,
    #  Entry Price, Shares, Capital Deployed,
    #  Shock Drop%, VIX Spike,
    #  Status, Current Price, Unrealized P/L,
    #  Date Closed, Exit Price, Realized P/L,
    #  Hold Days, Notes
    row = [
        trade_id, today, ticker, "LONG",
        round(price, 4), shares, round(capital, 2),
        shock_drop, vix_spike,
        "OPEN",
        round(price, 4), 0.0,
        "", "", "",
        0, "",
    ]
    ws.append(row)
    _style_data_row(ws, ws.max_row, "OPEN")
    _save(wb)
    return trade_id


def update_shock_prices():
    """Update current prices and unrealized P/L for all OPEN shock trades."""
    wb = _get_workbook()
    ws = _ensure_sheet(wb, "Shock", SHOCK_HEADERS)
    today = datetime.date.today().strftime("%Y-%m-%d")
    closed_ids = []

    # Column layout (1-indexed):
    #  1=Trade ID, 2=Date Opened, 3=Ticker, 4=Direction
    #  5=Entry Price, 6=Shares, 7=Capital Deployed
    #  8=Shock Drop%, 9=VIX Spike
    #  10=Status, 11=Current Price, 12=Unrealized P/L
    #  13=Date Closed, 14=Exit Price, 15=Realized P/L
    #  16=Hold Days, 17=Notes

    for row_idx in range(2, ws.max_row + 1):
        status = str(ws.cell(row=row_idx, column=10).value or "").strip().upper()
        if status == "OPEN":
            ticker    = ws.cell(row=row_idx, column=3).value
            entry     = float(ws.cell(row=row_idx, column=5).value or 0)
            shares    = float(ws.cell(row=row_idx, column=6).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            cur = _fetch_price(ticker)
            if cur is not None:
                ws.cell(row=row_idx, column=11, value=round(cur, 4))
            c   = cur if cur else entry
            pnl = round(shares * (c - entry), 2)
            ws.cell(row=row_idx, column=12, value=pnl)
            _color_pnl(ws, row_idx, 12, pnl)

            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                ws.cell(row=row_idx, column=16,
                        value=(datetime.date.today() - d_open).days)
            except ValueError:
                pass

        elif status == "CLOSE":
            ticker    = ws.cell(row=row_idx, column=3).value
            entry     = float(ws.cell(row=row_idx, column=5).value or 0)
            shares    = float(ws.cell(row=row_idx, column=6).value or 0)
            date_open = str(ws.cell(row=row_idx, column=2).value or today)

            exit_price = (_fetch_price(ticker)
                          or float(ws.cell(row=row_idx, column=11).value or entry))
            pnl = round(shares * (exit_price - entry), 2)

            ws.cell(row=row_idx, column=10, value="CLOSED")
            ws.cell(row=row_idx, column=11, value=round(exit_price, 4))
            ws.cell(row=row_idx, column=12, value="")
            ws.cell(row=row_idx, column=13, value=today)
            ws.cell(row=row_idx, column=14, value=round(exit_price, 4))
            ws.cell(row=row_idx, column=15, value=round(pnl, 2))
            _color_pnl(ws, row_idx, 15, pnl)
            _style_data_row(ws, row_idx, "CLOSED")

            try:
                d_open = datetime.datetime.strptime(date_open[:10], "%Y-%m-%d").date()
                ws.cell(row=row_idx, column=16,
                        value=(datetime.date.today() - d_open).days)
            except ValueError:
                pass

            closed_ids.append(ws.cell(row=row_idx, column=1).value)

    _save(wb)
    return closed_ids


# ── Summary sheet ────────────────────────────────────────────────────────────

def update_summary():
    """Refresh a Summary sheet with totals across all strategies."""
    wb = _get_workbook()

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)  # first tab

    ws.merge_cells("A1:D1")
    ws["A1"] = "Paper Trading Dashboard"
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")

    ws["A2"] = f"Last updated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, color="808080")

    row = 4
    for sheet_name in ["Pairs", "Momentum", "Bear", "Earnings", "Shock"]:
        if sheet_name not in wb.sheetnames:
            continue
        src = wb[sheet_name]
        # Determine column indices based on sheet
        if sheet_name == "Pairs":
            status_col, unreal_col, real_col = 11, 14, 18
        elif sheet_name == "Earnings":
            status_col, unreal_col, real_col = 11, 13, 16
        elif sheet_name == "Shock":
            status_col, unreal_col, real_col = 10, 12, 15
        else:
            status_col, unreal_col, real_col = 8, 10, 13

        open_count, closed_count = 0, 0
        total_unreal, total_real = 0.0, 0.0

        for r in range(2, src.max_row + 1):
            s = str(src.cell(row=r, column=status_col).value or "").strip().upper()
            if s == "OPEN":
                open_count += 1
                v = src.cell(row=r, column=unreal_col).value
                if v and str(v).strip():
                    total_unreal += float(v)
            elif s == "CLOSED":
                closed_count += 1
                v = src.cell(row=r, column=real_col).value
                if v and str(v).strip():
                    total_real += float(v)

        ws.cell(row=row, column=1, value=sheet_name).font = Font(bold=True, size=12)
        row += 1
        ws.cell(row=row, column=1, value="Open positions:")
        ws.cell(row=row, column=2, value=open_count)
        row += 1
        ws.cell(row=row, column=1, value="Closed trades:")
        ws.cell(row=row, column=2, value=closed_count)
        row += 1
        ws.cell(row=row, column=1, value="Unrealized P/L:")
        c = ws.cell(row=row, column=2, value=round(total_unreal, 2))
        c.number_format = _MONEY_FMT
        c.font = _GREEN_FONT if total_unreal >= 0 else _RED_FONT
        row += 1
        ws.cell(row=row, column=1, value="Realized P/L:")
        c = ws.cell(row=row, column=2, value=round(total_real, 2))
        c.number_format = _MONEY_FMT
        c.font = _GREEN_FONT if total_real >= 0 else _RED_FONT
        row += 2

    # Grand totals
    ws.cell(row=row, column=1, value="GRAND TOTAL").font = Font(bold=True, size=12)
    row += 1
    # Recalculate
    grand_unreal, grand_real = 0.0, 0.0
    for sheet_name in ["Pairs", "Momentum", "Bear", "Earnings"]:
        if sheet_name not in wb.sheetnames:
            continue
        src = wb[sheet_name]
        if sheet_name == "Pairs":
            status_col, unreal_col, real_col = 11, 14, 18
        elif sheet_name == "Earnings":
            status_col, unreal_col, real_col = 11, 13, 16
        else:
            status_col, unreal_col, real_col = 8, 10, 13
        for r in range(2, src.max_row + 1):
            s = str(src.cell(row=r, column=status_col).value or "").strip().upper()
            if s == "OPEN":
                v = src.cell(row=r, column=unreal_col).value
                if v and str(v).strip():
                    grand_unreal += float(v)
            elif s == "CLOSED":
                v = src.cell(row=r, column=real_col).value
                if v and str(v).strip():
                    grand_real += float(v)

    ws.cell(row=row, column=1, value="Total Unrealized:")
    c = ws.cell(row=row, column=2, value=round(grand_unreal, 2))
    c.number_format = _MONEY_FMT
    c.font = _GREEN_FONT if grand_unreal >= 0 else _RED_FONT
    row += 1
    ws.cell(row=row, column=1, value="Total Realized:")
    c = ws.cell(row=row, column=2, value=round(grand_real, 2))
    c.number_format = _MONEY_FMT
    c.font = _GREEN_FONT if grand_real >= 0 else _RED_FONT
    row += 1
    ws.cell(row=row, column=1, value="Combined:")
    combined = grand_unreal + grand_real
    c = ws.cell(row=row, column=2, value=round(combined, 2))
    c.number_format = _MONEY_FMT
    c.font = Font(bold=True, size=12, color="006100" if combined >= 0 else "9C0006")

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16

    _save(wb)


# ── Styling helpers ──────────────────────────────────────────────────────────

def _style_data_row(ws, row_idx, status):
    fill = _OPEN_FILL if status == "OPEN" else _CLOSED_FILL
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col).fill = fill


def _color_pnl(ws, row_idx, col, pnl):
    cell = ws.cell(row=row_idx, column=col)
    cell.number_format = _MONEY_FMT
    cell.font = _GREEN_FONT if pnl >= 0 else _RED_FONT


# ── CSV-to-Excel Sync ───────────────────────────────────────────────────────

def _read_csv_trades(csv_path: str) -> list:
    """Read a trades CSV into a list of dicts. Returns [] on failure."""
    if not os.path.exists(csv_path):
        return []
    try:
        with open(csv_path, "r", encoding="utf-8") as f:
            return list(csv.DictReader(f))
    except Exception:
        return []


def _sync_earn_csv(wb) -> int:
    """
    Sync earnings_trades.csv into the Earnings Excel sheet.
    - Closed CSV trades with OPEN Excel rows -> mark CLOSED with CSV exit data
    - CSV trades missing from Excel entirely -> append them
    Returns count of rows synced.
    """
    from config import EARN_TRADES_CSV
    csv_rows = _read_csv_trades(os.path.join(_SCRIPT_DIR, EARN_TRADES_CSV))
    if not csv_rows:
        return 0

    ws = _ensure_sheet(wb, "Earnings", EARN_HEADERS)
    synced = 0

    excel_ids = {}
    for row_idx in range(2, ws.max_row + 1):
        tid = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if tid:
            excel_ids[tid] = row_idx

    for cr in csv_rows:
        tid = cr.get("trade_id", "").strip()
        csv_status = cr.get("status", "").strip().lower()

        if tid in excel_ids:
            row_idx = excel_ids[tid]
            xl_status = str(ws.cell(row=row_idx, column=11).value or "").strip().upper()
            if xl_status == "OPEN" and csv_status == "closed":
                exit_price = float(cr.get("exit_price", 0) or 0)
                net_pnl = float(cr.get("net_pnl", 0) or 0)
                hold_days = cr.get("hold_days", "")
                date_close = cr.get("date_close", "")
                exit_reason = cr.get("exit_reason", "")

                ws.cell(row=row_idx, column=11, value="CLOSED")
                ws.cell(row=row_idx, column=12, value=round(exit_price, 4))
                ws.cell(row=row_idx, column=13, value="")
                ws.cell(row=row_idx, column=14, value=date_close)
                ws.cell(row=row_idx, column=15, value=round(exit_price, 4))
                ws.cell(row=row_idx, column=16, value=round(net_pnl, 2))
                _color_pnl(ws, row_idx, 16, net_pnl)
                if hold_days:
                    ws.cell(row=row_idx, column=17, value=int(hold_days))
                if exit_reason:
                    ws.cell(row=row_idx, column=18, value=exit_reason)
                _style_data_row(ws, row_idx, "CLOSED")
                synced += 1
        else:
            entry_price = float(cr.get("entry_price", 0) or 0)
            shares = float(cr.get("shares", 0) or 0)
            capital = float(cr.get("capital_deployed", 0) or 0)
            beat_rate = float(cr.get("beat_rate", 0) or 0) * 100
            is_closed = csv_status == "closed"

            row_data = [
                tid,
                cr.get("date_open", ""),
                cr.get("ticker", ""),
                cr.get("direction", "LONG"),
                round(entry_price, 4),
                round(shares, 4),
                round(capital, 2),
                cr.get("earnings_date", ""),
                round(beat_rate, 1),
                cr.get("eps_trend", ""),
                "CLOSED" if is_closed else "OPEN",
                round(float(cr.get("exit_price", 0) or 0), 4) if is_closed else round(entry_price, 4),
                "" if is_closed else 0.0,
                cr.get("date_close", "") if is_closed else "",
                round(float(cr.get("exit_price", 0) or 0), 4) if is_closed else "",
                round(float(cr.get("net_pnl", 0) or 0), 2) if is_closed else "",
                int(cr.get("hold_days", 0) or 0) if is_closed else 0,
                cr.get("exit_reason", "") if is_closed else "",
            ]
            ws.append(row_data)
            status = "CLOSED" if is_closed else "OPEN"
            _style_data_row(ws, ws.max_row, status)
            if is_closed:
                net_pnl = float(cr.get("net_pnl", 0) or 0)
                _color_pnl(ws, ws.max_row, 16, net_pnl)
            synced += 1

    return synced


def _sync_mom_csv(wb) -> int:
    """Sync momentum_trades.csv into the Momentum Excel sheet."""
    from config import MOM_LIVE_TRADES_CSV
    csv_rows = _read_csv_trades(os.path.join(_SCRIPT_DIR, MOM_LIVE_TRADES_CSV))
    if not csv_rows:
        return 0

    ws = _ensure_sheet(wb, "Momentum", MOM_HEADERS)
    synced = 0

    excel_ids = {}
    for row_idx in range(2, ws.max_row + 1):
        tid = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if tid:
            excel_ids[tid] = row_idx

    for cr in csv_rows:
        tid = cr.get("trade_id", "").strip()
        csv_status = cr.get("status", "").strip().lower()

        if tid in excel_ids:
            row_idx = excel_ids[tid]
            xl_status = str(ws.cell(row=row_idx, column=8).value or "").strip().upper()
            if xl_status == "OPEN" and csv_status == "closed":
                exit_price = float(cr.get("exit_price", 0) or 0)
                net_pnl = float(cr.get("net_pnl", 0) or 0)
                hold_days = cr.get("hold_days", "")
                date_close = cr.get("date_close", "")

                ws.cell(row=row_idx, column=8, value="CLOSED")
                ws.cell(row=row_idx, column=9, value=round(exit_price, 4))
                ws.cell(row=row_idx, column=10, value="")
                ws.cell(row=row_idx, column=11, value=date_close)
                ws.cell(row=row_idx, column=12, value=round(exit_price, 4))
                ws.cell(row=row_idx, column=13, value=round(net_pnl, 2))
                _color_pnl(ws, row_idx, 13, net_pnl)
                if hold_days:
                    ws.cell(row=row_idx, column=14, value=int(hold_days))
                _style_data_row(ws, row_idx, "CLOSED")
                synced += 1
        else:
            if csv_status != "closed":
                continue
            entry_price = float(cr.get("entry_price", 0) or 0)
            exit_price = float(cr.get("exit_price", 0) or 0)
            net_pnl = float(cr.get("net_pnl", 0) or 0)
            row_data = [
                tid, cr.get("date_open", ""), cr.get("ticker", ""),
                cr.get("direction", "LONG"),
                round(entry_price, 4),
                round(float(cr.get("shares", 0) or 0), 4),
                round(float(cr.get("capital_deployed", 0) or 0), 2),
                "CLOSED", round(exit_price, 4), "",
                cr.get("date_close", ""), round(exit_price, 4),
                round(net_pnl, 2),
                int(cr.get("hold_days", 0) or 0), "",
            ]
            ws.append(row_data)
            _style_data_row(ws, ws.max_row, "CLOSED")
            _color_pnl(ws, ws.max_row, 13, net_pnl)
            synced += 1

    return synced


def _sync_bear_csv(wb) -> int:
    """Sync bear_trades.csv into the Bear Excel sheet."""
    from config import BEAR_TRADES_CSV
    csv_rows = _read_csv_trades(os.path.join(_SCRIPT_DIR, BEAR_TRADES_CSV))
    if not csv_rows:
        return 0

    ws = _ensure_sheet(wb, "Bear", BEAR_HEADERS)
    synced = 0

    excel_ids = {}
    for row_idx in range(2, ws.max_row + 1):
        tid = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if tid:
            excel_ids[tid] = row_idx

    for cr in csv_rows:
        tid = cr.get("trade_id", "").strip()
        csv_status = cr.get("status", "").strip().lower()

        if tid in excel_ids:
            row_idx = excel_ids[tid]
            xl_status = str(ws.cell(row=row_idx, column=8).value or "").strip().upper()
            if xl_status == "OPEN" and csv_status == "closed":
                exit_price = float(cr.get("exit_price", 0) or 0)
                net_pnl = float(cr.get("net_pnl", 0) or 0)
                hold_days = cr.get("hold_days", "")
                date_close = cr.get("date_close", "")

                ws.cell(row=row_idx, column=8, value="CLOSED")
                ws.cell(row=row_idx, column=9, value=round(exit_price, 4))
                ws.cell(row=row_idx, column=10, value="")
                ws.cell(row=row_idx, column=11, value=date_close)
                ws.cell(row=row_idx, column=12, value=round(exit_price, 4))
                ws.cell(row=row_idx, column=13, value=round(net_pnl, 2))
                _color_pnl(ws, row_idx, 13, net_pnl)
                if hold_days:
                    ws.cell(row=row_idx, column=14, value=int(hold_days))
                _style_data_row(ws, row_idx, "CLOSED")
                synced += 1

    return synced


def _sync_shock_csv(wb) -> int:
    """Sync shock_trades.csv into the Shock Excel sheet."""
    from config import SHOCK_TRADES_CSV
    csv_rows = _read_csv_trades(os.path.join(_SCRIPT_DIR, SHOCK_TRADES_CSV))
    if not csv_rows:
        return 0

    ws = _ensure_sheet(wb, "Shock", SHOCK_HEADERS)
    synced = 0

    excel_ids = {}
    for row_idx in range(2, ws.max_row + 1):
        tid = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if tid:
            excel_ids[tid] = row_idx

    for cr in csv_rows:
        tid = cr.get("trade_id", "").strip()
        csv_status = cr.get("status", "").strip().lower()

        if tid in excel_ids:
            row_idx = excel_ids[tid]
            xl_status = str(ws.cell(row=row_idx, column=10).value or "").strip().upper()
            if xl_status == "OPEN" and csv_status == "closed":
                exit_price = float(cr.get("exit_price", 0) or 0)
                net_pnl = float(cr.get("net_pnl", 0) or 0)
                hold_days = cr.get("hold_days", "")
                date_close = cr.get("date_close", "")

                ws.cell(row=row_idx, column=10, value="CLOSED")
                ws.cell(row=row_idx, column=11, value=round(exit_price, 4))
                ws.cell(row=row_idx, column=12, value="")
                ws.cell(row=row_idx, column=13, value=date_close)
                ws.cell(row=row_idx, column=14, value=round(exit_price, 4))
                ws.cell(row=row_idx, column=15, value=round(net_pnl, 2))
                _color_pnl(ws, row_idx, 15, net_pnl)
                if hold_days:
                    ws.cell(row=row_idx, column=16, value=int(hold_days))
                _style_data_row(ws, row_idx, "CLOSED")
                synced += 1

    return synced


def _sync_pairs_csv(wb) -> int:
    """Sync live_trades.csv into the Pairs Excel sheet."""
    from config import LIVE_TRADES_CSV
    csv_rows = _read_csv_trades(os.path.join(_SCRIPT_DIR, LIVE_TRADES_CSV))
    if not csv_rows:
        return 0

    ws = _ensure_sheet(wb, "Pairs", PAIRS_HEADERS)
    synced = 0

    excel_ids = {}
    for row_idx in range(2, ws.max_row + 1):
        tid = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if tid:
            excel_ids[tid] = row_idx

    for cr in csv_rows:
        tid = cr.get("trade_id", "").strip()
        csv_status = cr.get("status", "").strip().lower()

        if tid in excel_ids:
            row_idx = excel_ids[tid]
            xl_status = str(ws.cell(row=row_idx, column=11).value or "").strip().upper()
            if xl_status == "OPEN" and csv_status == "closed":
                exit_a = float(cr.get("price_a_exit", 0) or 0)
                exit_b = float(cr.get("price_b_exit", 0) or 0)
                net_pnl = float(cr.get("net_pnl", 0) or 0)
                hold_days = cr.get("hold_days", "")
                date_close = cr.get("date_close", "")

                ws.cell(row=row_idx, column=11, value="CLOSED")
                ws.cell(row=row_idx, column=12, value=round(exit_a, 4))
                ws.cell(row=row_idx, column=13, value=round(exit_b, 4))
                ws.cell(row=row_idx, column=14, value="")
                ws.cell(row=row_idx, column=15, value=date_close)
                ws.cell(row=row_idx, column=16, value=round(exit_a, 4))
                ws.cell(row=row_idx, column=17, value=round(exit_b, 4))
                ws.cell(row=row_idx, column=18, value=round(net_pnl, 2))
                _color_pnl(ws, row_idx, 18, net_pnl)
                if hold_days:
                    ws.cell(row=row_idx, column=19, value=int(hold_days))
                _style_data_row(ws, row_idx, "CLOSED")
                synced += 1

    return synced


# ── Master update (called once per daily run) ────────────────────────────────

def update_all():
    """
    Update all open positions with current prices, process any CLOSE requests,
    and refresh the summary dashboard. Called from run_system.py.
    """
    print("  Excel Tracker: updating paper_trades.xlsx ...")

    # Sync CSV closures/missing trades into Excel before price updates
    try:
        wb = _get_workbook()
        total_synced = 0
        total_synced += _sync_pairs_csv(wb)
        total_synced += _sync_mom_csv(wb)
        total_synced += _sync_bear_csv(wb)
        total_synced += _sync_earn_csv(wb)
        total_synced += _sync_shock_csv(wb)
        if total_synced > 0:
            _save(wb)
            print(f"    [CSV Sync] {total_synced} trades synced from CSV")
    except Exception as exc:
        print(f"    [CSV Sync] Warning: {exc}")

    closed = []

    c = update_pairs_prices()
    if c:
        closed.extend(c)
        for tid in c:
            print(f"    [Pairs] CLOSED: {tid}")

    c = update_momentum_prices()
    if c:
        closed.extend(c)
        for tid in c:
            print(f"    [Momentum] CLOSED: {tid}")

    c = update_bear_prices()
    if c:
        closed.extend(c)
        for tid in c:
            print(f"    [Bear] CLOSED: {tid}")

    c = update_earn_prices()
    if c:
        closed.extend(c)
        for tid in c:
            print(f"    [Earnings] CLOSED: {tid}")

    c = update_shock_prices()
    if c:
        closed.extend(c)
        for tid in c:
            print(f"    [Shock] CLOSED: {tid}")

    update_summary()

    # Count open positions
    wb = _get_workbook()
    total_open = 0
    for sheet_name in ["Pairs", "Momentum", "Bear", "Earnings", "Shock"]:
        if sheet_name not in wb.sheetnames:
            continue
        src = wb[sheet_name]
        if sheet_name in ("Pairs", "Earnings"):
            status_col = 11
        elif sheet_name == "Shock":
            status_col = 10
        else:
            status_col = 8
        for r in range(2, src.max_row + 1):
            s = str(src.cell(row=r, column=status_col).value or "").strip().upper()
            if s == "OPEN":
                total_open += 1

    print(f"    Open positions: {total_open}  |  "
          f"Newly closed: {len(closed)}")
    print(f"    File: {EXCEL_PATH}\n")
    return closed
